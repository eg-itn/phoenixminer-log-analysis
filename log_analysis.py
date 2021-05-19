import glob
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
import pandas as pd


# ログファイル処理
def read_log():
    # ログのタイムスタンプをExcelの形式に変化する関数
    def convert_timestamp(timestamp:str):
        timestamp = timestamp[:4] + '/' + timestamp[5:7] + '/' + timestamp[8:10] + ' ' + timestamp[11:]
        return timestamp

    # ログ読み込み
    log_list = glob.iglob('log/**.txt')  # ログファイルリスト
    power_list = []  # 消費電力リスト
    speed_list = []  # 採掘速度リスト

    # ログファイルごとに処理
    for log in log_list:
        print(log)
        with open(log, 'r') as f:
            while True:
                line = f.readline()
                if not line:
                    break
                elif line == '\n':
                    continue

                # タイムスタンプ
                timestamp = line[:23]
                timestamp = convert_timestamp(timestamp)

                line = f.readline()

                # 消費電力取得
                if line[:10] == 'GPUs power':
                    end = line.find(' W')
                    # print(timestamp + ' Power ' + line[12:18])
                    power_list.append(timestamp + ',' + line[12:end])  # 電力行はタイムスタンプがないので1行前のを使用
                # 採掘速度取得
                elif line[35:50] == 'Effective speed':
                    timestamp = line[:23]
                    timestamp = convert_timestamp(timestamp)
                    # print(timestamp + ' EffSpped ' + line[52:57])
                    speed_list.append(timestamp + ',,' + line[52:57])

    return power_list, speed_list


# csv作成
def write_csv(outcsv, power_list, speed_list):
    """
    :param outcsv: 出力先csv
    :param power_list: 消費電力リスト
    :param speed_list: 採掘速度リスト
    """
    with open(outcsv, 'w') as fw:
        fw.write(',power,speed\n')
        for line in power_list:
            fw.write(line)
            fw.write('\n')
        for line in speed_list:
            fw.write(line)
            fw.write('\n')


# Excel作成
def write_excel(power_list, speed_list, outcsv, report):
    """
    :param power_list: 消費電力リスト
    :param speed_list: 採掘速度リスト
    :param outcsv: write_csvで保存したデータ
    :param report: 出力先Excel
    :return: csvを読み込んだDataframe
    """
    df = pd.read_csv(outcsv)  # csv読み込み
    df.iloc[:,0] = pd.to_datetime(df.iloc[:,0])  # タイムスタンプをdatetime形式に変更
    with pd.ExcelWriter(report) as writer:  # Excelにデータ出力
        df.to_excel(writer, index=False)

    # Excel読み込み
    wb = openpyxl.load_workbook(report)
    ws = wb.active

    # グラフのサイズ、位置を指定
    chart = ScatterChart()
    chart.width = 40
    chart.height = 20
    pos_x = 4
    pos_y = 1

    # y軸データ
    values_power = Reference(ws, min_row=1, max_row=len(power_list), min_col=2, max_col=2)
    values_speed = Reference(ws, min_row=len(power_list)+2, max_row=ws.max_row, min_col=3, max_col=3)
    # values = Reference(ws, min_row=1, max_row=ws.max_row, min_col=2, max_col=3)
    chart.legend.legendPos = 'b'

    # x軸データ
    x_axis_power = Reference(ws, min_row=2, max_row=len(power_list), min_col=1, max_col=1)
    x_axis_speed = Reference(ws, min_row=len(power_list)+2, max_row=ws.max_row, min_col=1, max_col=1)
    # x_axis = Reference(ws, min_row=2, max_row=ws.max_row, min_col=1, max_col=1)
    # chart.set_categories(x_axis)

    # 電力データをチャートに追加
    series_power = Series(values_power, x_axis_power, title_from_data=True)
    chart.series.append(series_power)
    # 速度データをチャートに追加
    series_speed = Series(values_speed, x_axis_speed, title_from_data=True)
    chart.series.append(series_speed)

    # チャート描画
    ws.add_chart(chart, ws.cell(row=pos_y, column=pos_x).coordinate)

    wb.save(report)

    return df


def calculate_profit(df):
    pass


def main():
    power_list, speed_list = read_log()
    outcsv = 'output.csv'
    report = 'output.xlsx'
    write_csv(outcsv, power_list, speed_list)
    df = write_excel(power_list, speed_list, outcsv, report)



if __name__ == '__main__':
    main()
